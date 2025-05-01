from io import BytesIO

import openpyxl
from openpyxl.cell.cell import Cell


def adjust_excel_file(
    exported_data: bytes,
    row_width: int | None = None,
    row_height: int | None = None,
) -> bytes:
    # Load an excel file from exported data.
    excel_file = openpyxl.load_workbook(BytesIO(exported_data))

    # Adjust worksheet names.
    for worksheet in excel_file.worksheets:
        worksheet.title = "Data"

    # Set column width height and for all rows with any value.
    for worksheet in excel_file.worksheets:
        for column in worksheet.iter_cols(
            min_row=worksheet.min_row, max_row=worksheet.max_row
        ):
            for row in column:
                if isinstance(row, Cell):
                    if row_width:
                        worksheet.column_dimensions[row.column_letter].width = row_width
                    if row_height:
                        worksheet.row_dimensions[row.row].height = row_height

    # Convert an excel file to bytes.
    buffer = BytesIO()
    excel_file.save(buffer)
    return buffer.getvalue()
