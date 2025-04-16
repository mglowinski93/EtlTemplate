from io import BytesIO

import openpyxl
from openpyxl.cell.cell import Cell


def adjust_excel_file(
    title: str,
    exported_data: bytes,
    row_width: int | None = None,
    row_height: int | None = None,
) -> bytes:
    # Load an excel file from exported data.
    excel_file = openpyxl.load_workbook(BytesIO(exported_data))

    # Set column width height and for all rows with any value.
    for worksheet in excel_file.worksheets:
        for column in worksheet.iter_cols(
            min_row=worksheet.min_row, max_row=worksheet.max_row
        ):
            for row in column:
                if isinstance(row, Cell):
                    if row_width:
                        worksheet.column_dimensions[row.column_letter].width = row_width
                    if row_height:
                        worksheet.row_dimensions[row.row].height = row_height

    # Add title row.
    for worksheet in excel_file.worksheets:
        worksheet.insert_rows(idx=1, amount=1)
        worksheet.merge_cells(
            f"A1:{worksheet.cell(row=1, column=worksheet.max_column).coordinate}"
        )
        worksheet["A1"] = title

    # Convert an excel file to bytes.
    buffer = BytesIO()
    excel_file.save(buffer)
    return buffer.getvalue()
